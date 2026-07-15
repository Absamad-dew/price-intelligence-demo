import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from price_demo import run


ROOT = Path(__file__).resolve().parents[1]
EXPECTED_SHEETS = [
    "normalized_offers",
    "best_prices",
    "review_queue",
    "quality_metrics",
    "imports",
]
FORMULA_ERROR_MARKERS = ("#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A")


class OpenpyxlInteroperabilityTests(unittest.TestCase):
    def test_generated_workbook_has_valid_structure_and_formulas(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory)
            run(ROOT / "samples", output, ROOT / "config.json")
            workbook_path = output / "price_intelligence.xlsx"

            workbook = load_workbook(workbook_path, data_only=False, read_only=False)
            try:
                self.assertEqual(workbook.sheetnames, EXPECTED_SHEETS)

                for worksheet in workbook.worksheets:
                    self.assertEqual(worksheet.freeze_panes, "A2")
                    self.assertIsNotNone(worksheet.auto_filter.ref)
                    self.assertGreaterEqual(worksheet.max_row, 2)
                    self.assertGreaterEqual(worksheet.max_column, 1)

                    headers = [cell.value for cell in worksheet[1]]
                    self.assertTrue(all(isinstance(header, str) and header for header in headers))
                    self.assertEqual(len(headers), len(set(headers)))

                    for row in worksheet.iter_rows():
                        for cell in row:
                            if cell.data_type != "f":
                                continue
                            self.assertIsInstance(cell.value, str)
                            self.assertTrue(cell.value.startswith("="))
                            self.assertFalse(
                                any(marker in cell.value.upper() for marker in FORMULA_ERROR_MARKERS),
                                f"broken formula marker in {worksheet.title}!{cell.coordinate}: {cell.value}",
                            )

                best_prices = workbook["best_prices"]
                best_headers = [cell.value for cell in best_prices[1]]
                canonical_id_column = best_headers.index("canonical_id") + 1
                canonical_ids = {
                    best_prices.cell(row=row, column=canonical_id_column).value
                    for row in range(2, best_prices.max_row + 1)
                }
                self.assertEqual(
                    canonical_ids,
                    {"bolt-m8x30", "cable-vvg-3x2.5", "lithium-grease"},
                )
            finally:
                workbook.close()

            values_workbook = load_workbook(workbook_path, data_only=True, read_only=True)
            try:
                error_cells = [
                    f"{worksheet.title}!{cell.coordinate}"
                    for worksheet in values_workbook.worksheets
                    for row in worksheet.iter_rows()
                    for cell in row
                    if cell.data_type == "e"
                ]
                self.assertEqual(error_cells, [])
            finally:
                values_workbook.close()


if __name__ == "__main__":
    unittest.main()
